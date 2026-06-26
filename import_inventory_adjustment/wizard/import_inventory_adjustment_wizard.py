from odoo import models, fields, _
from odoo.exceptions import ValidationError, UserError
import base64
from io import BytesIO
from datetime import datetime, date
import openpyxl


class ImportInventoryAdjustmentWizard(models.TransientModel):
    _name = "import.inventory.adjustment.wizard"
    _description = "Import Inventory Adjustment"

    location_ids = fields.Many2many("stock.location", string="Locations", domain=[("usage", "=", "internal")])
    file = fields.Binary(required=True)
    filename = fields.Char()
    result_file = fields.Binary(readonly=True)
    result_filename = fields.Char(readonly=True)

    def action_import(self):
        wb = openpyxl.load_workbook(
            BytesIO(base64.b64decode(self.file)),
            data_only=True
        )
        sheet = wb.active

        headers = [(c.value or "").strip() for c in sheet[1]]

        LOCATION_COL = 0   # Column A
        LOT_COL      = 1   # Column B
        PRODUCT_COL  = 2   # Column C
        QTY_COL      = 3   # Column D (optional)    
        ACC_DATE_COL     = 4   # Column E (optional)

        # def col(name):
        #     return headers.index(name) if name in headers else None

        # location_col = col("Location")
        # product_col = col("Product")
        # qty_col = col("Counted Quantity")
        # lot_col = col("Lot/Serial Number")
        # account_date_col = col("Accounting date")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            location_col = row[LOCATION_COL]
            product_col  = row[PRODUCT_COL]
            qty_col = row[QTY_COL]
            lot_col = row[LOT_COL] if len(row) > LOT_COL else None
            account_date_col = row[ACC_DATE_COL] if len(row) > ACC_DATE_COL else None

        # ---- RESULT XLSX ----
        result_wb = openpyxl.Workbook()
        result_ws = result_wb.active
        result_ws.append([
            "Location", "Product", "Qty", "Lot",
            "Accounting Date", "Status", "Message"
        ])

        for row in sheet.iter_rows(min_row=2):
            status = "Skipped"
            message = ""

            try:
                value = row[PRODUCT_COL].value or ""
                code = value.split("]")[0].replace("[", "").strip()

                product = self.env["product.product"].search(
                    [("default_code", "=", code)], limit=1
                )
                if not product:
                    raise ValidationError("Product not found")

                location_name = row[LOCATION_COL].value
                if not location_name:
                    raise ValidationError("Location empty")

                location = self.env["stock.location"].search(
                    [("complete_name", "=", location_name)], limit=1
                )
                if not location:
                    raise ValidationError("Location not found")

                # lot_id = False
                # if lot_col is not None and row[LOT_COL].value:
                #     lot = self.env["stock.lot"].search([
                #         ("name", "=", row[LOT_COL].value),
                #         ("product_id", "=", product.id)
                #     ], limit=1)
                #     if not lot:
                #         raise ValidationError("Lot not found")
                #     lot_id = lot.id

                lot_id = False
                lot_name = row[LOT_COL].value if LOT_COL < len(row) else False
                if lot_name:
                    lot = self.env["stock.lot"].search([
                        ("name", "=", str(lot_name)),
                        ("product_id", "=", product.id),
                        ("company_id", "=", self.env.company.id)
                    ], limit=1)

                    # 🔹 CREATE LOT IF NOT FOUND
                    if not lot:
                        lot = self.env["stock.lot"].create({
                            "name": str(lot_name),
                            "product_id": product.id,
                            "company_id": self.env.company.id,
                        })

                    lot_id = lot.id

                qty = row[QTY_COL].value or 0
                cell_value = row[ACC_DATE_COL].value if account_date_col is not None else None
                accounting_date = False

                if cell_value:
                    if isinstance(cell_value, datetime):
                        accounting_date = cell_value.date()
                    elif isinstance(cell_value, date):
                        accounting_date = cell_value
                    elif isinstance(cell_value, str):
                        cell_value = cell_value.strip()

                        for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S"):
                            try:
                                accounting_date = datetime.strptime(cell_value, fmt).date()
                                break
                            except ValueError:
                                continue

                domain = [
                    ("location_id", "=", location.id),
                    ("product_id", "=", product.id),
                    ("lot_id", "=", lot_id or False),
                ]

                quant = self.env["stock.quant"].search(domain, limit=1)

                if quant:
                    quant.write({
                        "inventory_quantity": qty,
                        "inventory_quantity_set": True,
                    })
                    status = "Updated"
                    message = "Quant updated"
                elif not quant and (not self.location_ids or location in self.location_ids):

                    stock_quatnt_id = self.env["stock.quant"].create({
                        "product_id": product.id,
                        "location_id": location.id,
                        "inventory_quantity": qty,
                        "accounting_date": accounting_date,
                        "inventory_quantity_set": True,
                        "lot_id": lot_id or False,

                    })
                    # stock_quatnt_id._compute_inventory_diff_quantity()
                    # self.env.cr.commit()
                    status = "Created"
                    message = "Quant created"

            except Exception as e:
                message = str(e)

            # ---- WRITE RESULT ROW ----
            result_ws.append([
                row[LOCATION_COL].value or "",
                row[PRODUCT_COL].value or "",
                row[QTY_COL].value or 0,
                row[LOT_COL].value if len(row) > LOT_COL else "",
                row[ACC_DATE_COL].value if len(row) > ACC_DATE_COL else "",
                status,
                message,
            ])

        # ---- EXPORT RESULT XLSX ----
        buffer = BytesIO()
        result_wb.save(buffer)
        buffer.seek(0)

        self.result_file = base64.b64encode(buffer.read())
        self.result_filename = "inventory_import_result.xlsx"

        return {
            "type": "ir.actions.act_url",
            "url": f"/web/content/?model={self._name}&id={self.id}"
                   f"&field=result_file&filename={self.result_filename}&download=true",
            "target": "self",
        }
